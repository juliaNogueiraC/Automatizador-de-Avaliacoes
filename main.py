import openpyxl
from fpdf import FPDF
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


# Função para carregar a lista de alunos a partir de um arquivo TXT
def carregar_lista_alunos(arquivo_txt):
    with open(arquivo_txt, 'r', encoding='utf-8') as f:
        alunos = [linha.strip() for linha in f.readlines()]
    return alunos

# Função para atualizar o modelo de avaliação com o nome do aluno
def atualizar_modelo(nome_aluno, modelo_excel, nova_pasta_excel, celula_destino):
    wb = openpyxl.load_workbook(modelo_excel)
    ws = wb.active

    # Inserir o nome do aluno com o prefixo "NOME COMPLETO:"
    ws[celula_destino].value = f"NOME COMPLETO: {nome_aluno}"

    novo_arquivo_excel = os.path.join(nova_pasta_excel, f'{nome_aluno}.xlsx')
    wb.save(novo_arquivo_excel)
    return novo_arquivo_excel

# Função para converter uma planilha Excel em PDF preservando a formatação

def excel_para_pdf(arquivo_excel, arquivo_pdf):
    wb = openpyxl.load_workbook(arquivo_excel)
    ws = wb.active

    pdf = canvas.Canvas(arquivo_pdf, pagesize=A4)
    y = 750  # Posição inicial na página

    for row in ws.iter_rows():
        for cell in row:
            pdf.drawString(50, y, str(cell.value))
            y -= 12  # Move para a próxima linha

    pdf.save()

# O restante do código permanece o mesmo...


def main():
    arquivo_txt = 'lista_alunos.txt'
    modelo_excel = 'modelo_avaliacao.xlsx'
    nova_pasta_excel = 'avaliacoes_excel2'
    nova_pasta_pdf = 'avaliacoes_pdf2'

    # Criar pastas se não existirem
    os.makedirs(nova_pasta_excel, exist_ok=True)
    os.makedirs(nova_pasta_pdf, exist_ok=True)

    celula_destino = 'A2'  # Especifique a célula de destino para o nome do aluno

    alunos = carregar_lista_alunos(arquivo_txt)
    print(f'Lista de alunos carregada: {alunos}')

    for aluno in alunos:
        print(f'Processando avaliação para {aluno}')
        novo_arquivo_excel = atualizar_modelo(aluno, modelo_excel, nova_pasta_excel, celula_destino)
        print(f'Arquivo Excel atualizado para {aluno}: {novo_arquivo_excel}')
        arquivo_pdf = os.path.join(nova_pasta_pdf, f'{aluno}.pdf')
        excel_para_pdf(novo_arquivo_excel, arquivo_pdf)
        print(f'Avaliação de {aluno} salva como PDF na pasta {nova_pasta_pdf}.')

if __name__ == '__main__':
    main()
